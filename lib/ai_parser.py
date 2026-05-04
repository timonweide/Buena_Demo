"""
Buena Piloto — AI Parser (BWA → Buena Standard P&L)

Dual-Mode:
- Wenn ANTHROPIC_API_KEY gesetzt → echter Claude API Call
- Sonst → Mock Mode mit deterministischen Outputs (für Demo ohne Internet)

Workflow:
1. User uploadet BWA (Excel oder PDF)
2. extract_text_from_bwa() liefert strukturierten Text der GuV-Zeilen
3. parse_bwa_with_ai() ruft Claude API ODER Mock-Logik
4. Output: List[MappingResult] mit source_label, mapped_to (Buena-Schema), confidence, override_value
"""
import os
import json
import io
from typing import Any
from openpyxl import load_workbook


# ──────────────────────────────────────────────────────────────────────────
# Buena Standard P&L Schema
# ──────────────────────────────────────────────────────────────────────────
BUENA_PNL_SCHEMA = {
    "Revenue": [
        "WEG-Verwaltung",
        "Mietverwaltung",
        "Sondereigentumsverwaltung",
        "Sondervergütungen",
        "Other Revenue",
    ],
    "Direct Costs": [
        "Personnel - Property Managers",
        "Personnel - Accounting",
        "Personnel - Other Ops",
    ],
    "Operating Expenses": [
        "Software & IT",
        "Office & Rent",
        "Marketing & Sales",
        "Insurance",
        "Travel",
        "Other OpEx",
    ],
    "D&A": [
        "Depreciation Tangibles",
        "Depreciation Intangibles",
    ],
    "Financial": [
        "Interest Income",
        "Interest Expense",
    ],
    "Tax": [
        "Corporate Tax",
        "Trade Tax",
    ],
    "_subtotals": ["Gross Profit", "EBITDA", "EBIT", "EBT", "Net Income"],
}


def all_buena_categories() -> list[str]:
    """Liefert alle gültigen Buena-Kategorien für Mapping (inkl. _Skip)."""
    cats = []
    for category, items in BUENA_PNL_SCHEMA.items():
        if category.startswith("_"):
            continue
        for item in items:
            cats.append(f"{category} → {item}")
    cats.append("_Skip / Subtotal")
    return cats


# ──────────────────────────────────────────────────────────────────────────
# BWA Extraction
# ──────────────────────────────────────────────────────────────────────────
def extract_lines_from_excel(file_bytes: bytes) -> list[dict]:
    """
    Liest BWA Excel und extrahiert Zeilen mit Konto, Bezeichnung, Wert.
    Robust gegen verschiedene Layouts.
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    lines = []
    for row in ws.iter_rows(values_only=True):
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue

        cells = [str(c).strip() if c is not None else "" for c in row]

        konto = ""
        label = ""
        amount = None

        for c in row:
            if c is None:
                continue
            if isinstance(c, (int, float)):
                if amount is None:
                    amount = float(c)
            elif isinstance(c, str):
                cs = c.strip()
                if not cs:
                    continue
                if cs.isdigit() and 3 <= len(cs) <= 6 and not konto:
                    konto = cs
                elif len(cs) > 5 and not label:
                    label = cs
                elif not label:
                    label = cs

        if label and amount is not None:
            lines.append({"konto": konto, "label": label, "amount": amount})

    return lines


def extract_lines_from_pdf(file_bytes: bytes) -> list[dict]:
    """
    Extrahiert GuV-Zeilen aus einer PDF-BWA via pdfplumber.
    Funktioniert für text-basierte PDFs (nicht gescannte Bilder).
    """
    try:
        import pdfplumber
    except ImportError:
        return []

    lines = []
    import re

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for raw_line in text.split("\n"):
                raw_line = raw_line.strip()
                if not raw_line or len(raw_line) < 5:
                    continue

                # Suche nach Zeilen mit einer Zahl am Ende (GuV-Zeilen)
                # Typisches Format: "Bezeichnung   1.234,56" oder "4400 Bezeichnung 1.234,56"
                # Unterstützt DE-Format (Punkt als Tausender, Komma als Dezimal)
                # und EN-Format (Komma als Tausender, Punkt als Dezimal)
                amount_pat = re.search(
                    r"([+-]?[\d]{1,3}(?:[.,]?\d{3})*(?:[.,]\d{1,2})?)(?:\s*€?)?\s*$",
                    raw_line,
                )
                if not amount_pat:
                    continue

                amount_str = amount_pat.group(1)
                # Normalize: DE-Format → float
                try:
                    # "1.234,56" → "1234.56"
                    if "," in amount_str and "." in amount_str:
                        # Thousands sep = ".", decimal sep = ","
                        amount_str_norm = amount_str.replace(".", "").replace(",", ".")
                    elif "," in amount_str:
                        amount_str_norm = amount_str.replace(",", ".")
                    else:
                        amount_str_norm = amount_str
                    amount = float(amount_str_norm)
                except ValueError:
                    continue

                # Label = everything before the number
                label_part = raw_line[: amount_pat.start()].strip()
                if not label_part or len(label_part) < 3:
                    continue

                # Try to extract Konto-Nr from start of label
                konto = ""
                label = label_part
                konto_match = re.match(r"^(\d{3,6})\s+(.+)$", label_part)
                if konto_match:
                    konto = konto_match.group(1)
                    label = konto_match.group(2).strip()

                if label and amount != 0:
                    lines.append({"konto": konto, "label": label, "amount": amount})

    return lines


def extract_lines_from_bwa(filename: str, file_bytes: bytes) -> list[dict]:
    """Router je nach Dateityp — Excel oder PDF."""
    name_lower = filename.lower()
    if name_lower.endswith((".xlsx", ".xlsm")):
        return extract_lines_from_excel(file_bytes)
    elif name_lower.endswith(".pdf"):
        return extract_lines_from_pdf(file_bytes)
    else:
        return []


# ──────────────────────────────────────────────────────────────────────────
# Mock Mode — Deterministic Mapping (für Demo ohne API-Key)
# ──────────────────────────────────────────────────────────────────────────
MOCK_KEYWORD_MAP = [
    # (keyword tuple, target category, confidence)
    (("erlös", "weg", "verwaltung"), "Revenue → WEG-Verwaltung", 0.92),
    (("erlös", "sev"), "Revenue → Sondereigentumsverwaltung", 0.88),
    (("erlös", "sondereigentum"), "Revenue → Sondereigentumsverwaltung", 0.95),
    (("erlös", "miet"), "Revenue → Mietverwaltung", 0.95),
    (("sondervergütung",), "Revenue → Sondervergütungen", 0.94),
    (("mahn", "verzug"), "Revenue → Other Revenue", 0.78),
    (("erlös",), "Revenue → Other Revenue", 0.65),

    (("subverwalter", "fremdleistung"), "Operating Expenses → Other OpEx", 0.72),
    (("edv", "software", "lizenz"), "Operating Expenses → Software & IT", 0.93),

    (("löhne", "verwalter"), "Direct Costs → Personnel - Property Managers", 0.96),
    (("gehälter", "buchhaltung"), "Direct Costs → Personnel - Accounting", 0.96),
    (("gehälter", "geschäftsleitung"), "Direct Costs → Personnel - Other Ops", 0.85),
    (("sozialvers",), "Direct Costs → Personnel - Other Ops", 0.78),
    (("berufsgenossenschaft",), "Direct Costs → Personnel - Other Ops", 0.85),
    (("pensionskasse",), "Direct Costs → Personnel - Other Ops", 0.85),
    (("personalnebenkosten", "weiterbildung"), "Direct Costs → Personnel - Other Ops", 0.80),

    (("miete", "geschäftsräume"), "Operating Expenses → Office & Rent", 0.95),
    (("nebenkosten", "geschäftsräume"), "Operating Expenses → Office & Rent", 0.92),
    (("reinigung", "hausmeister"), "Operating Expenses → Office & Rent", 0.85),
    (("versicherung",), "Operating Expenses → Insurance", 0.92),
    (("reisekosten",), "Operating Expenses → Travel", 0.95),
    (("werbung", "marketing"), "Operating Expenses → Marketing & Sales", 0.94),
    (("repräsentation", "bewirtung"), "Operating Expenses → Marketing & Sales", 0.70),
    (("bürobedarf",), "Operating Expenses → Other OpEx", 0.85),
    (("telefon", "internet"), "Operating Expenses → Software & IT", 0.78),
    (("rechts", "beratung"), "Operating Expenses → Other OpEx", 0.82),
    (("steuerberatung", "jahresabschluss"), "Operating Expenses → Other OpEx", 0.85),
    (("buchführung",), "Operating Expenses → Other OpEx", 0.85),
    (("sonstige", "betrieblich"), "Operating Expenses → Other OpEx", 0.65),

    (("abschreibung", "sachanlagen"), "D&A → Depreciation Tangibles", 0.95),
    (("abschreibung", "immateriell"), "D&A → Depreciation Intangibles", 0.95),

    (("zinsertrag",), "Financial → Interest Income", 0.95),
    (("zinsaufwand",), "Financial → Interest Expense", 0.95),

    (("körperschaftsteuer", "solz"), "Tax → Corporate Tax", 0.95),
    (("gewerbesteuer",), "Tax → Trade Tax", 0.95),
]


def mock_map_line(line: dict) -> dict:
    """Deterministisches Mapping basierend auf Keyword-Match."""
    label_lower = line["label"].lower()

    # Subtotals / Header lines = skip
    skip_indicators = ["σ ", "ebitda", "ebit", "ebt", "jahresüberschuss",
                       "umsatzerlöse", "betriebsergebnis", "net income"]
    if any(s in label_lower for s in skip_indicators):
        return {
            "source_label": line["label"],
            "source_konto": line["konto"],
            "source_amount": line["amount"],
            "mapped_to": "_Skip / Subtotal",
            "confidence": 0.99,
            "rationale": "Erkannt als Zwischensumme oder Header — wird in Buena-Schema separat berechnet",
        }

    # Best keyword match (longest matching tuple first)
    best_match = None
    best_score = -1
    for keywords, target_cat, conf in MOCK_KEYWORD_MAP:
        if all(kw in label_lower for kw in keywords):
            score = len(keywords)
            if score > best_score:
                best_score = score
                best_match = (target_cat, conf, keywords)

    if best_match:
        target_cat, conf, keywords = best_match
        return {
            "source_label": line["label"],
            "source_konto": line["konto"],
            "source_amount": line["amount"],
            "mapped_to": target_cat,
            "confidence": conf,
            "rationale": f"Keyword-Match: {', '.join(keywords)}",
        }

    return {
        "source_label": line["label"],
        "source_konto": line["konto"],
        "source_amount": line["amount"],
        "mapped_to": "Operating Expenses → Other OpEx",
        "confidence": 0.40,
        "rationale": "Kein klarer Keyword-Match — Default auf Other OpEx, manuelle Prüfung empfohlen",
    }


# ──────────────────────────────────────────────────────────────────────────
# Real Claude API Mode
# ──────────────────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """Du bist ein Financial Analyst spezialisiert auf deutsche Hausverwaltungen \
und DATEV-SKR04-Buchhaltungssysteme. Deine Aufgabe: Mappe BWA-Zeilen (deutscher \
Kontenrahmen) auf das Buena Standard P&L Schema.

Buena Standard P&L Kategorien:
- Revenue → WEG-Verwaltung (recurring, Verwaltung von Wohnungseigentümergemeinschaften)
- Revenue → Mietverwaltung (recurring, Verwaltung Mietobjekte)
- Revenue → Sondereigentumsverwaltung (SEV, recurring)
- Revenue → Sondervergütungen (project-based, z.B. WEG-Versammlungen)
- Revenue → Other Revenue (Mahnwesen etc.)
- Direct Costs → Personnel - Property Managers (Verwalter)
- Direct Costs → Personnel - Accounting (Buchhaltung)
- Direct Costs → Personnel - Other Ops (Geschäftsleitung, SVB, Berufsgenossenschaft, Pensionskasse, Sonstiges Personal)
- Operating Expenses → Software & IT (EDV, Telefon, Internet)
- Operating Expenses → Office & Rent (Miete, Nebenkosten, Reinigung)
- Operating Expenses → Marketing & Sales (Werbung, Repräsentation)
- Operating Expenses → Insurance (Versicherungen)
- Operating Expenses → Travel (Reisekosten)
- Operating Expenses → Other OpEx (Rechtsberatung, Steuerberatung, Sonstiges)
- D&A → Depreciation Tangibles (Sachanlagen)
- D&A → Depreciation Intangibles (Immaterielle VG)
- Financial → Interest Income (Zinserträge)
- Financial → Interest Expense (Zinsaufwand)
- Tax → Corporate Tax (Körperschaftsteuer + SolZ)
- Tax → Trade Tax (Gewerbesteuer)
- _Skip / Subtotal (für Zwischensummen, EBITDA-Zeilen, Header)

Antwort-Format: NUR valides JSON, kein Markdown, kein Kommentar drumherum:
{"mappings": [{"source_label": "...", "mapped_to": "Revenue → ...", "confidence": 0.92, "rationale": "..."}]}

Jede Mapping-Entscheidung: confidence zwischen 0.0 und 1.0. Bei <0.7 = manuelle Review nötig.
Rationale: kurz, max 12 Wörter."""


def claude_map_lines(lines: list[dict]) -> list[dict]:
    """Echter Claude API Call. Wirft Exception wenn kein API-Key oder API fehlschlägt."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY nicht gesetzt")

    import anthropic
    client = anthropic.Anthropic(api_key=api_key)

    # Build user message
    line_descriptions = "\n".join(
        f"{i+1}. Konto {l['konto'] or '—'}: {l['label']} ({l['amount']:,.0f} €)"
        for i, l in enumerate(lines)
    )
    user_msg = f"""Bitte mappe folgende BWA-Zeilen auf das Buena Standard P&L Schema:

{line_descriptions}

Antworte NUR mit JSON, exakt im Format aus dem System-Prompt."""

    response = client.messages.create(
        model="claude-opus-4-7",
        max_tokens=4000,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_msg}],
    )

    raw = response.content[0].text.strip()
    # Strip code fences if any
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    raw = raw.strip()

    parsed = json.loads(raw)
    mappings = parsed.get("mappings", [])

    # Reattach amounts and konto numbers
    enriched = []
    for i, m in enumerate(mappings):
        if i >= len(lines):
            break
        enriched.append({
            "source_label": m.get("source_label", lines[i]["label"]),
            "source_konto": lines[i]["konto"],
            "source_amount": lines[i]["amount"],
            "mapped_to": m.get("mapped_to", "Operating Expenses → Other OpEx"),
            "confidence": float(m.get("confidence", 0.5)),
            "rationale": m.get("rationale", "—"),
        })
    return enriched


# ──────────────────────────────────────────────────────────────────────────
# Public Entry Point
# ──────────────────────────────────────────────────────────────────────────
def parse_bwa(filename: str, file_bytes: bytes, force_mock: bool = False) -> dict:
    """
    Hauptfunktion. Liefert:
        {
            "mode": "claude" | "mock",
            "lines_extracted": int,
            "mappings": [...],
            "error": str | None
        }
    """
    lines = extract_lines_from_bwa(filename, file_bytes)

    if not lines:
        return {
            "mode": "error",
            "lines_extracted": 0,
            "mappings": [],
            "error": "Keine GuV-Zeilen extrahiert. Format unterstützt: Excel (.xlsx).",
        }

    use_real_api = bool(os.environ.get("ANTHROPIC_API_KEY")) and not force_mock

    if use_real_api:
        try:
            mappings = claude_map_lines(lines)
            return {"mode": "claude", "lines_extracted": len(lines), "mappings": mappings, "error": None}
        except Exception as e:
            # Fallback to mock on any API failure
            mappings = [mock_map_line(l) for l in lines]
            return {
                "mode": "mock",
                "lines_extracted": len(lines),
                "mappings": mappings,
                "error": f"Claude API fehlgeschlagen, Fallback auf Mock: {str(e)[:100]}",
            }
    else:
        mappings = [mock_map_line(l) for l in lines]
        return {"mode": "mock", "lines_extracted": len(lines), "mappings": mappings, "error": None}


# ──────────────────────────────────────────────────────────────────────────
# Aggregation: Mappings → Buena Standard P&L
# ──────────────────────────────────────────────────────────────────────────
def aggregate_to_buena_pnl(mappings: list[dict]) -> dict:
    """
    Aggregiert Mappings zu Buena Standard P&L mit Subtotals.
    Berücksichtigt manuelle Overrides (key 'override_to' falls gesetzt).
    """
    totals = {}
    for m in mappings:
        target = m.get("override_to") or m["mapped_to"]
        if target == "_Skip / Subtotal":
            continue
        amount = m["source_amount"] or 0
        totals[target] = totals.get(target, 0) + amount

    # Build structured output
    structured = {category: {} for category in BUENA_PNL_SCHEMA if not category.startswith("_")}

    for category, items in BUENA_PNL_SCHEMA.items():
        if category.startswith("_"):
            continue
        for item in items:
            key = f"{category} → {item}"
            structured[category][item] = totals.get(key, 0)

    # Subtotals
    revenue_total = sum(structured["Revenue"].values())
    direct_costs_total = sum(structured["Direct Costs"].values())  # already negative
    opex_total = sum(structured["Operating Expenses"].values())
    da_total = sum(structured["D&A"].values())
    fin_total = sum(structured["Financial"].values())
    tax_total = sum(structured["Tax"].values())

    gross_profit = revenue_total + direct_costs_total
    ebitda = gross_profit + opex_total
    ebit = ebitda + da_total
    ebt = ebit + fin_total
    net_income = ebt + tax_total

    return {
        "structured": structured,
        "subtotals": {
            "Revenue Total": revenue_total,
            "Gross Profit": gross_profit,
            "EBITDA": ebitda,
            "EBIT": ebit,
            "EBT": ebt,
            "Net Income": net_income,
        },
        "ebitda_margin_pct": (ebitda / revenue_total * 100) if revenue_total > 0 else 0,
    }
